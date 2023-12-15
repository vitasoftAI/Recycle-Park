import os, torch, random, torchvision.transforms as tfs, numpy as np, xlsxwriter, openpyxl, matplotlib.pyplot as plt, timm
from openpyxl.styles import PatternFill
from collections import OrderedDict as OD
from PIL import Image, ImageFont
from tqdm import tqdm

def predict_classes(qry_fms_all, pos_fms_all, cls_names, im_lbls, num_top_stop, top_k):
    
    cos = torch.nn.CosineSimilarity(dim = 1, eps = 1e-6)
    for idx, qry in enumerate(qry_fms_all):
        if idx == num_top_stop: break
        print(f"\n{idx + 1}번 이미지 결과 (정답      -> {cls_names[im_lbls[idx]]}):\n")
        vals, inds = torch.topk(cos(qry, pos_fms_all), k = top_k)
        
        for ii, (val, ind) in enumerate(zip(vals, inds)):
            print(f"{val.item():.3f} 확률로 top{ii+1} 파트번호 -> {cls_names[im_lbls[ind.item()]]}")
            
def predict_classes_new(model, qry_fms_all, im_lbls, test_dl, model_name, data_name, device, cls_names, num_top_stop, top_k, lang, save_path = "test_dbs"):
    
    os.makedirs(save_path, exist_ok = True)
    db_path_qry = f"{save_path}/{model_name}_{data_name}_qry.pth"
    db_path_ims = f"{save_path}/{model_name}_{data_name}_ims.pth"
    cos = torch.nn.CosineSimilarity(dim = 1, eps = 1e-6)
    
    if os.path.isfile(db_path_qry) and os.path.isfile(db_path_ims):
        if lang == "en": print("Found saved database! Loading...")
        elif lang == "ko": print("저장된 데이터베이스가 있습니다! 로드중입니다...")
        ims_fms_all, ims_all = torch.load(db_path_qry), torch.load(db_path_ims)
    else:
        # ims_fms_all, ims_all = [], []

        if lang == "en": print("Obtaining embeddings...")
        elif lang == "ko": print("피처 추출하는중....")
        paths, top5_inds_all, qry_paths = [], [], []
        batch_count, error = 0 , 0
        # try:
        for i, (ims, im_paths) in tqdm(enumerate(test_dl)):
            # if batch_count == 20: break
            ims = ims.to(device)

            with torch.no_grad():
                ims_fms = get_fm(model.forward_features(ims))
                # ims_fms_all.extend(ims_fms)
                # ims_all.extend(ims.to(device))

#         ims_fms_all = torch.stack(ims_fms_all)
#         ims_all = torch.stack(ims_all)

#         torch.save(ims_fms_all, db_path_qry)
#         torch.save(ims_all, db_path_ims)

    # if lang == "en": print("Embeddings are obtained!")
    # elif lang == "ko": print("피처 추출 완료!")

                for idx, qry in enumerate(ims_fms):
                    print(f"\n{(batch_count * 64) + (idx + 1)}번 이미지 결과:\n")

                    # print(f"\n{idx + 1}번 이미지 결과 (정답      -> {cls_names[im_lbls[idx]]}):\n")
                    vals, inds = torch.topk(cos(qry, qry_fms_all), k = 200)
                    top5_vals, top5_inds = [], []

                    for i, (val, ind) in enumerate(zip(vals, inds)):
                        if len(top5_inds) == top_k: break
                        pred_class = cls_names[im_lbls[ind.item()].item()]
                        if pred_class not in top5_inds: top5_inds.append(pred_class); top5_vals.append(val)
                        else: continue

                    top5_inds_all.append(top5_inds)
                    qry_paths.append(im_paths[idx])
                    for ii, (val, ind) in enumerate(zip(top5_vals, top5_inds)):
                        print(f"{val.item():.3f} 확률로 top{ii+1} 파트번호 -> {ind}")
            batch_count += 1
            
        # except: print("Skipping batch with errors..."); error += 1
    # print(f"There are {error} batches with errors!")
    
    excel_summary(data_type = data_name, paths = qry_paths, top5s = top5_inds_all)

def create_dbs(model, test_dl, model_name, data_name, device, lang, save_path = "saved_dbs"):
    
    os.makedirs(save_path, exist_ok = True)
    db_path_pos = f"{save_path}/{model_name}_{data_name}_pos.pth"
    db_path_qry = f"{save_path}/{model_name}_{data_name}_qry.pth"
    db_path_ims = f"{save_path}/{model_name}_{data_name}_ims.pth"
    db_path_lbls = f"{save_path}/{model_name}_{data_name}_lbls.pth"
    
    if os.path.isfile(db_path_qry) and os.path.isfile(db_path_pos) and os.path.isfile(db_path_ims):
        qry_fms_all = torch.load(db_path_qry)
        pos_fms_all = torch.load(db_path_pos)
        ims_all = torch.load(db_path_ims)
        im_lbls = torch.load(db_path_lbls)
        if lang == "en": print("Embeddings are obtained!")
        elif lang == "ko": print("피처 추출 완료!")
    else:
        qry_fms_all, pos_fms_all, ims_all, im_lbls = [], [], [], []

        if lang == "en": print("Obtaining embeddings...")
        elif lang == "ko": print("피처 추출하는중....")
        for idx, batch in enumerate(tqdm(test_dl)):

            # if idx == 100: break
            qry_ims, pos_ims, qry_lbls = batch["qry_im"], batch["pos_im"], batch["qry_im_lbl"]
            qry_ims = qry_ims.to(device)

            with torch.no_grad():
                qry_fms = get_fm(model.forward_features(qry_ims))
                # pos_fms = get_fm(model.forward_features(pos_ims))
                qry_fms_all.extend(qry_fms)
                im_lbls.extend(qry_lbls)
                # pos_fms_all.extend(pos_fms)
                # ims_all.extend(qry_ims)
                

        qry_fms_all = torch.stack(qry_fms_all)
        # pos_fms_all = torch.FloatTensor(pos_fms_all)
        # ims_all = torch.stack(ims_all)
        im_lbls = torch.stack(im_lbls)
        
        torch.save(qry_fms_all, db_path_qry)
        # torch.save(pos_fms_all, db_path_pos)
        # torch.save(ims_all, db_path_ims)
        torch.save(im_lbls, db_path_lbls)
        
        if lang == "en": print("Embeddings are obtained!")
        elif lang == "ko": print("피처 추출 완료!")

    # return ims_all, qry_fms_all, pos_fms_all, im_lbls
    return qry_fms_all, im_lbls

def get_model(model_name, n_cls, device, saved_model_path, lang):
    
    model = timm.create_model(model_name, num_classes = n_cls)
    model.to(device)
    model.load_state_dict(get_state_dict(saved_model_path), strict = True)
    if lang == "en":
        print(f"{model_name} model trained weights are successfully loaded!")
    elif lang == "ko":
        print(f"{model_name} 학습된 AI모델의 가중치가 성공적으로 로드되었습니다!")
    
    return model

def makedirs(path): os.makedirs(path, exist_ok = True)
    
def get_state_dict(checkpoint_path):
    
    checkpoint = torch.load(checkpoint_path, map_location = "cpu")
    new_state_dict = OD()
    for k, v in checkpoint["state_dict"].items():
        name = k.replace("model.", "") # remove `model.`
        new_state_dict[name] = v
    return new_state_dict

def tn2np(t, t_type = "rgb", with_norm = False):
    
    assert t_type in ["rgb", "gray"], "Rasm RGB yoki grayscale ekanligini aniqlashtirib bering."
    
    gray_tfs = tfs.Compose([tfs.Normalize(mean = [ 0.], std = [1/0.5]), tfs.Normalize(mean = [-0.5], std = [1])])
    rgb_tfs = tfs.Compose([tfs.Normalize(mean = [ 0., 0., 0. ], std = [ 1/0.229, 1/0.224, 1/0.225 ]), tfs.Normalize(mean = [ -0.485, -0.456, -0.406 ], std = [ 1., 1., 1. ])])
    
    invTrans = gray_tfs if t_type == "gray" else rgb_tfs 
    
    return (invTrans(t) * 255).detach().cpu().permute(1,2,0).numpy().astype(np.uint8) if with_norm else (t * 255).detach().cpu().permute(1,2,0).numpy().astype(np.uint8)

def predict(m, path, tfs, cls_names):
    
    fontpath = "SpoqaHanSansNeo-Light.ttf"
    font = ImageFont.truetype(fontpath, 200)
    im = Image.open(path)
    im.save(path)
    
    return im, cls_names[int(torch.max(m(tfs(im).unsqueeze(0)).data, 1)[1])]

# @st.cache_data
def load_model(model_name, num_classes, checkpoint_path): 
    
    """
    
    This function gets several parameters and loads a classification model.
    
    Parameters:
    
        model_name      - name of a model from timm library, str;
        num_classes     - number of classes in the dataset, int;
        checkpoint_path - path to the trained model, str;
        
    Output:
    
        m              - a model with pretrained weights and in an evaluation mode, torch model object;
    
    """
    
    m = timm.create_model(model_name, num_classes = num_classes)
    m.load_state_dict(checkpoint_path)
    
    return m.eval()

def get_state_dict(checkpoint_path):
    
    checkpoint = torch.load(checkpoint_path)
    new_state_dict = OD()
    for k, v in checkpoint["state_dict"].items():
        name = k.replace("model.", "") # remove `model.`
        new_state_dict[name] = v
    return new_state_dict

def get_fm(fm):
        
        """
        
        This function gets feature map with size (bs, fm_shape, 7, 7)
        applies average pooling and returns feature map with shape (bs, fm_shape).
        
        Parameter:
        
            fm - feature map, tensor.
        
        Output:
        
            fm - reshaped feature map, tensor.
        
        """
        
        pool = torch.nn.AvgPool2d((fm.shape[2],fm.shape[3]))
        
        return torch.reshape(pool(fm), (-1, fm.shape[1]))
    
def visualize(ds, num_ims, row, cmap = None, cls_names = None):
    
    plt.figure(figsize = (20, 10))
    indekslar = [random.randint(0, len(ds) - 1) for _ in range(num_ims)]
    for idx, indeks in enumerate(indekslar):
        
        im, gt = ds[indeks]
        # Start plot
        plt.subplot(row, num_ims // row, idx + 1)
        if cmap:
            plt.imshow(tensor_2_im(im), cmap='gray')
        else:
            plt.imshow(tensor_2_im(im))
        plt.axis('off')
        if cls_names is not None:
            plt.title(f"GT -> {cls_names[str(gt)]}")
        else:
            plt.title(f"GT -> {gt}")
            
def data_tekshirish(ds):
    
    data = ds[0]    
    print(f"Dataning birinchi elementining turi: {type(data[0])}")
    print(f"Dataning ikkinchi elementining turi: {type(data[1])}")
    print(f"Dataning birinchi elementining hajmi: {(data[0]).shape}")
    print(f"Dataning birinchi elementidagi piksel qiymatlari: {np.unique(np.array(data[0]))}")
    print(f"Dataning ikkinchi elementi: {data[1]}")
    

def tensor_2_im(t, t_type = "rgb", with_norm = False):
    
    assert t_type in ["rgb", "gray"], "Rasm RGB yoki grayscale ekanligini aniqlashtirib bering."
    
    gray_tfs = tfs.Compose([tfs.Normalize(mean = [ 0.], std = [1/0.5]), tfs.Normalize(mean = [-0.5], std = [1])])
    rgb_tfs = tfs.Compose([tfs.Normalize(mean = [ 0., 0., 0. ], std = [ 1/0.229, 1/0.224, 1/0.225 ]), tfs.Normalize(mean = [ -0.485, -0.456, -0.406 ], std = [ 1., 1., 1. ])])
    
    invTrans = gray_tfs if t_type == "gray" else rgb_tfs 
    
    return (invTrans(t) * 255).detach().cpu().permute(1,2,0).numpy().astype(np.uint8) if with_norm else (t * 255).detach().cpu().permute(1,2,0).numpy().astype(np.uint8)

def parametrlar_soni(model): 
    for name, param in model.named_parameters():
        print(f"{name} parametrida {param.numel()} ta parametr bor.")
    print(f"Modelning umumiy parametrlar soni -> {sum(param.numel() for param in model.parameters() if param.requires_grad)} ta.")
    
def inference(model, device, test_dl, num_ims, row, cls_names = None):
    
    preds, images, lbls = [], [], []
    for idx, data in enumerate(test_dl):
        im, gt = data
        im, gt = im.to(device), gt.to(device)
        _, pred = torch.max(model(im), dim = 1)
        images.append(im)
        preds.append(pred.item())
        lbls.append(gt.item())
    
    plt.figure(figsize = (20, 10))
    indekslar = [random.randint(0, len(images) - 1) for _ in range(num_ims)]
    for idx, indeks in enumerate(indekslar):
        
        im = images[indeks].squeeze()
        # Start plot
        plt.subplot(row, num_ims // row, idx + 1)
        plt.imshow(tensor_2_im(im), cmap='gray')
        plt.axis('off')
        if cls_names is not None: plt.title(f"GT -> {cls_names[str(lbls[indeks])]} ; Prediction -> {cls_names[str(preds[indeks])]}", color=("green" if {cls_names[str(lbls[indeks])]} == {cls_names[str(preds[indeks])]} else "red"))
        else: plt.title(f"GT -> {gt} ; Prediction -> {pred}")
        
def make_xlsx(data_type, di): 
    
    workbook = xlsxwriter.Workbook(f"excel_files/{data_type}.xlsx")
    colors = ['0000FF00', '00FF0000']
    fillers = []

    for color in colors:
        temp = PatternFill(patternType = "solid", fgColor = color)
        fillers.append(temp)

    worksheet = workbook.add_worksheet()
    worksheet.write(f'A1', f'{data_type}')
    worksheet.write(f'B1', '클래스별로')
    worksheet.write(f'C1', '이미지 수')

    for idx, (key, value) in enumerate(di.items()):

        worksheet.write(f'A{idx + 2}', f'{key}')
        worksheet.write(f'B{idx + 2}', f'{value}')
    workbook.close()

    wb = openpyxl.load_workbook(f"excel_files/{data_type}.xlsx")
    ws = wb['Sheet1']

    for idx, (key, value) in enumerate(di.items()):
        ws[f"B{idx + 2}"].fill = fillers[0] if value >= 30 else fillers[1]
    wb.save(f"excel_files/{data_type}.xlsx")
    
    
def excel_summary(data_type, paths, top5s): 
    
    workbook = xlsxwriter.Workbook(f"excel_files/{data_type}_top5.xlsx")
    colors = ['0000FF00', '00FF0000']
    fillers = []

    for color in colors:
        temp = PatternFill(patternType = "solid", fgColor = color)
        fillers.append(temp)

    worksheet = workbook.add_worksheet()
    worksheet.write(f'A1', f'{data_type}_ims_paths')
    worksheet.write(f'B1', 'top1')
    worksheet.write(f'C1', 'top2')
    worksheet.write(f'D1', 'top3')
    worksheet.write(f'E1', 'top4')
    worksheet.write(f'F1', 'top5')

    for idx, (path, top5) in enumerate(zip(paths, top5s)):
        if len(top5) == 5:
            worksheet.write(f'A{idx + 2}', f'{path}')
            worksheet.write(f'B{idx + 2}', f'{top5[0]}')
            worksheet.write(f'C{idx + 2}', f'{top5[1]}')
            worksheet.write(f'D{idx + 2}', f'{top5[2]}')
            worksheet.write(f'E{idx + 2}', f'{top5[3]}')
            worksheet.write(f'F{idx + 2}', f'{top5[4]}')
        elif len(top5) == 4:
            worksheet.write(f'A{idx + 2}', f'{path}')
            worksheet.write(f'B{idx + 2}', f'{top5[0]}')
            worksheet.write(f'C{idx + 2}', f'{top5[1]}')
            worksheet.write(f'D{idx + 2}', f'{top5[2]}')
            worksheet.write(f'E{idx + 2}', f'{top5[3]}')
        elif len(top5) == 3:
            worksheet.write(f'A{idx + 2}', f'{path}')
            worksheet.write(f'B{idx + 2}', f'{top5[0]}')
            worksheet.write(f'C{idx + 2}', f'{top5[1]}')
            worksheet.write(f'D{idx + 2}', f'{top5[2]}')
        else:
            worksheet.write(f'A{idx + 2}', f'{path}')
            worksheet.write(f'B{idx + 2}', f'{top5[0]}')
            worksheet.write(f'C{idx + 2}', f'{top5[1]}')
    workbook.close()

def calculator(all_preds, all_gts, cls_names, metric):
    
    di = {}
    pr_scores = metric(all_preds, all_gts)
    for idx, score in enumerate(pr_scores):
        di[cls_names[idx]] = f"{score.item():.3f}"
        
    return di  

def make_metric_xlsx(data_type, metric_type, di): 
    
    workbook = xlsxwriter.Workbook(f"excel_files/{data_type}_{metric_type}.xlsx")
    colors = ['0000FF00', '00FF0000']
    fillers = []

    for color in colors:
        temp = PatternFill(patternType = "solid", fgColor = color)
        fillers.append(temp)

    worksheet = workbook.add_worksheet()
    worksheet.write(f'A1', f'파트번호')
    worksheet.write(f'B1', f'{metric_type} 평가 점수')

    for idx, (key, value) in enumerate(di.items()):

        worksheet.write(f'A{idx + 2}', f'{key}')
        worksheet.write(f'B{idx + 2}', f'{value}')
    workbook.close()

    wb = openpyxl.load_workbook(f"excel_files/{data_type}_{metric_type}.xlsx")
    ws = wb['Sheet1']

    for idx, (key, value) in enumerate(di.items()):
        ws[f"B{idx + 2}"].fill = fillers[0] if float(value) >= 0.9 else fillers[1]
    wb.save(f"excel_files/{data_type}_{metric_type}.xlsx")